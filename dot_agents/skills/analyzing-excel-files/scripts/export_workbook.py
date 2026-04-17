#!/usr/bin/env -S uv run --script
# /// script
# requires-python = ">=3.11"
# dependencies = ["openpyxl>=3.1.5"]
# ///

from __future__ import annotations

import argparse
import csv
import json
import re
from pathlib import Path

from openpyxl import load_workbook


def slugify(value: str) -> str:
    return re.sub(r"[^a-zA-Z0-9]+", "_", value.strip()).strip("_").lower() or "sheet"


def main() -> None:
    parser = argparse.ArgumentParser(description="Export workbook sheets to CSV and JSON fixtures.")
    parser.add_argument("workbook", type=Path)
    parser.add_argument("output_dir", type=Path)
    args = parser.parse_args()

    workbook_path = args.workbook.expanduser().resolve()
    output_dir = args.output_dir.expanduser().resolve()
    output_dir.mkdir(parents=True, exist_ok=True)

    workbook = load_workbook(
        workbook_path,
        data_only=True,
        keep_vba=workbook_path.suffix.lower() == ".xlsm",
    )

    manifest = {
        "workbook": str(workbook_path),
        "sheet_names": workbook.sheetnames,
        "sheets": [],
    }

    for ws in workbook.worksheets:
        slug = slugify(ws.title)
        rows = [[cell.value for cell in row] for row in ws.iter_rows()]

        csv_path = output_dir / f"{slug}.csv"
        with csv_path.open("w", encoding="utf-8", newline="") as handle:
            writer = csv.writer(handle)
            for row in rows:
                writer.writerow(["" if value is None else value for value in row])

        json_path = output_dir / f"{slug}.json"
        json_path.write_text(
            json.dumps(
                {
                    "title": ws.title,
                    "max_row": ws.max_row,
                    "max_column": ws.max_column,
                    "hidden_rows": [
                        row_idx
                        for row_idx in range(1, ws.max_row + 1)
                        if ws.row_dimensions[row_idx].hidden
                    ],
                    "hidden_columns": [
                        col_letter
                        for col_letter, dim in ws.column_dimensions.items()
                        if dim.hidden
                    ],
                    "freeze_panes": None if ws.freeze_panes is None else str(ws.freeze_panes),
                    "rows": rows,
                },
                indent=2,
                ensure_ascii=False,
                default=str,
            )
            + "\n"
        )

        manifest["sheets"].append(
            {
                "title": ws.title,
                "slug": slug,
                "csv": csv_path.name,
                "json": json_path.name,
            }
        )

    (output_dir / "manifest.json").write_text(json.dumps(manifest, indent=2) + "\n")


if __name__ == "__main__":
    main()
