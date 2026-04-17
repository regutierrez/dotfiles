#!/usr/bin/env -S uv run --script
# /// script
# requires-python = ">=3.11"
# dependencies = ["openpyxl>=3.1.5"]
# ///

from __future__ import annotations

import argparse
import json
from pathlib import Path

from openpyxl import load_workbook


def build_summary(path: Path) -> dict:
    workbook = load_workbook(path, data_only=False, keep_vba=path.suffix.lower() == ".xlsm")
    return {
        "workbook": str(path),
        "sheet_names": workbook.sheetnames,
        "defined_names": list(workbook.defined_names.keys()),
        "sheets": [
            {
                "title": ws.title,
                "max_row": ws.max_row,
                "max_column": ws.max_column,
                "formula_count": sum(
                    1
                    for row in ws.iter_rows()
                    for cell in row
                    if isinstance(cell.value, str) and cell.value.startswith("=")
                ),
                "hidden_row_count": sum(
                    1 for i in range(1, ws.max_row + 1) if ws.row_dimensions[i].hidden
                ),
                "hidden_column_count": sum(1 for dim in ws.column_dimensions.values() if dim.hidden),
                "merged_range_count": len(ws.merged_cells.ranges),
                "freeze_panes": None if ws.freeze_panes is None else str(ws.freeze_panes),
            }
            for ws in workbook.worksheets
        ],
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Print workbook structure as JSON.")
    parser.add_argument("workbook", type=Path)
    args = parser.parse_args()
    print(json.dumps(build_summary(args.workbook.expanduser().resolve()), indent=2))


if __name__ == "__main__":
    main()
