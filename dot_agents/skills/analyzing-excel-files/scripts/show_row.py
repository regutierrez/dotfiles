#!/usr/bin/env -S uv run --script
# /// script
# requires-python = ">=3.11"
# dependencies = ["openpyxl>=3.1.5"]
# ///

from __future__ import annotations

import argparse
import json
from pathlib import Path

from openpyxl import load_workbook


def row_values(path: Path, sheet: str, row_number: int, data_only: bool) -> list[dict[str, object]]:
    workbook = load_workbook(path, data_only=data_only, keep_vba=path.suffix.lower() == ".xlsm")
    ws = workbook[sheet]
    max_col = ws.max_column
    header_row = max(1, row_number - 1)
    result: list[dict[str, object]] = []
    for col_idx in range(1, max_col + 1):
        value = ws.cell(row_number, col_idx).value
        header = ws.cell(header_row, col_idx).value
        if value is None and header is None:
            continue
        result.append({
            "column": col_idx,
            "header": header,
            "value": value,
        })
    return result


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Show a representative row in formula mode and cached-value mode."
    )
    parser.add_argument("workbook", type=Path)
    parser.add_argument("sheet")
    parser.add_argument("row", type=int)
    args = parser.parse_args()

    path = args.workbook.expanduser().resolve()
    payload = {
        "workbook": str(path),
        "sheet": args.sheet,
        "row": args.row,
        "formula_view": row_values(path, args.sheet, args.row, data_only=False),
        "cached_value_view": row_values(path, args.sheet, args.row, data_only=True),
    }
    print(json.dumps(payload, indent=2, ensure_ascii=False, default=str))


if __name__ == "__main__":
    main()
